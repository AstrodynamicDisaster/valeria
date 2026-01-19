#!/usr/bin/env python3
"""
Modelo 190 generator.

Implements Modelo 190 file generation for Spanish IRPF annual summary.
Focuses on clave A and clave L subclaves 01/05/24/25/26 by default.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from sqlalchemy import Date, cast, func, or_, select
from sqlalchemy.orm import Session

from core.database import get_session
from core.models import Client, ClientLocation, Employee, EmployeePeriod, Payroll, PayrollLine


_DECIMAL_ZERO = Decimal("0.00")
_EUR_QUANT = Decimal("0.01")


@dataclass(frozen=True)
class ConceptMappingRule:
    clave: str
    subclave: Optional[str]
    concepts: List[str]


@dataclass
class MappingConfig:
    concepts_to_clave_subclave: List[ConceptMappingRule] = field(default_factory=list)
    ss_tax_concepts: List[str] = field(default_factory=list)
    allowed_claves: Sequence[str] = field(default_factory=lambda: ["A", "L"])
    allowed_subclaves_by_clave: Dict[str, Sequence[str]] = field(
        default_factory=lambda: {"L": ["01", "05", "24", "25", "26"]}
    )
    gastos_deducibles_allocation: str = "clave_a"  # "clave_a" or "proportional"
    allow_gastos_deducibles_claves: Sequence[str] = field(
        default_factory=lambda: ["A", "L.05", "L.10", "L.27"]
    )
    include_in_kind: bool = False
    default_provincia: str = "98"

    def build_concept_map(self) -> Dict[str, Tuple[str, Optional[str]]]:
        concept_map: Dict[str, Tuple[str, Optional[str]]] = {}
        for entry in self.concepts_to_clave_subclave:
            for raw_concept in entry.concepts:
                key = normalize_concept_key(raw_concept)
                if not key:
                    continue
                if key in concept_map:
                    raise ValueError(f"Concept mapping conflict for '{raw_concept}'.")
                concept_map[key] = (entry.clave, entry.subclave)
        return concept_map


@dataclass
class AdditionalData:
    anio_nacimiento: Optional[int] = None
    situacion_familiar: Optional[int] = None
    nif_conyuge: Optional[str] = None
    discapacidad: Optional[int] = None
    contrato_relacion: Optional[int] = None
    movilidad_geografica: Optional[bool] = None
    gastos_deducibles: Optional[Decimal] = None
    comunicacion_prestamo_vivienda: Optional[bool] = None


@dataclass
class IngresoDinerario:
    base: Decimal
    retencion: Decimal = _DECIMAL_ZERO
    sign: Optional[str] = None


@dataclass
class IngresoEspecie:
    base: Decimal
    ingreso_cuenta_efectuado: Decimal = _DECIMAL_ZERO
    ingreso_cuenta_repercutido: Decimal = _DECIMAL_ZERO
    sign: Optional[str] = None


@dataclass
class Incapacidad:
    dineraria: Optional[IngresoDinerario] = None
    especie: Optional[IngresoEspecie] = None


@dataclass
class ForalSplit:
    hacienda_estatal: Decimal = _DECIMAL_ZERO
    navarra: Decimal = _DECIMAL_ZERO
    araba: Decimal = _DECIMAL_ZERO
    gipuzkoa: Decimal = _DECIMAL_ZERO
    bizkaia: Decimal = _DECIMAL_ZERO


@dataclass
class PerceptorRecordInput:
    nif_perceptor: str
    nombre_perceptor: str
    provincia: str
    clave: str
    subclave: Optional[str] = None
    nif_representante: Optional[str] = None
    ejercicio_devengo: Optional[int] = None
    ceuta_melilla_flag: Optional[int] = None
    dinerario_no_incapacidad: Optional[IngresoDinerario] = None
    especie_no_incapacidad: Optional[IngresoEspecie] = None
    incapacidad: Optional[Incapacidad] = None
    complemento_ayuda_infancia: Optional[int] = None
    excesos_acciones_emergentes: Optional[int] = None
    foral_split: Optional[ForalSplit] = None
    datos_adicionales: Optional[AdditionalData] = None


@dataclass
class Declarant190:
    ejercicio: int
    modelo: str
    nif: str
    nombre_razon_social: str
    tipo_declaracion: str
    percepciones: List[PerceptorRecordInput]
    contacto_telefono: Optional[str] = None
    contacto_nombre: Optional[str] = None
    email_contacto: Optional[str] = None
    numero_identificativo: Optional[str] = None
    id_declaracion_anterior: Optional[str] = None


@dataclass
class Aggregation:
    percepcion_no_it: Decimal = _DECIMAL_ZERO
    percepcion_it: Decimal = _DECIMAL_ZERO
    retencion_no_it: Decimal = _DECIMAL_ZERO
    retencion_it: Decimal = _DECIMAL_ZERO
    gastos_deducibles: Decimal = _DECIMAL_ZERO
    devengo_base_no_it: Decimal = _DECIMAL_ZERO


@dataclass
class EmployeeInfo:
    employee_id: int
    nif: str
    first_name: str
    last_name: str
    last_name2: Optional[str]
    birth_date: Optional[date]
    address: Optional[str]


@dataclass(frozen=True)
class GroupKey:
    employee_id: int
    clave: str
    subclave: Optional[str]


def _to_decimal(value: Optional[Decimal | int | float | str]) -> Decimal:
    if value is None:
        return _DECIMAL_ZERO
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _quantize_eur(value: Decimal) -> Decimal:
    return value.quantize(_EUR_QUANT, rounding=ROUND_HALF_UP)


def _as_float(value: Optional[Decimal]) -> float:
    if value is None:
        return 0.0
    return float(_quantize_eur(_to_decimal(value)))


_ACCENT_TRANSLATION = str.maketrans(
    {
        "\u00c1": "A",
        "\u00c9": "E",
        "\u00cd": "I",
        "\u00d3": "O",
        "\u00da": "U",
        "\u00c0": "A",
        "\u00c8": "E",
        "\u00cc": "I",
        "\u00d2": "O",
        "\u00d9": "U",
        "\u00c2": "A",
        "\u00ca": "E",
        "\u00ce": "I",
        "\u00d4": "O",
        "\u00db": "U",
        "\u00c4": "A",
        "\u00cb": "E",
        "\u00cf": "I",
        "\u00d6": "O",
        "\u00dc": "U",
    }
)


_ALLOWED_TEXT_RE = re.compile(r"[^A-Z0-9 \u00d1\u00c7]")


def normalize_text(value: Optional[str]) -> str:
    if not value:
        return ""
    text = value.upper()
    text = text.translate(_ACCENT_TRANSLATION)
    text = re.sub(r"[\t\n\r]+", " ", text)
    text = _ALLOWED_TEXT_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_concept_key(concept: Optional[str]) -> str:
    if not concept:
        return ""
    return normalize_text(concept)


def fmt_alpha(value: Optional[str], length: int) -> str:
    normalized = normalize_text(value)
    if len(normalized) > length:
        normalized = normalized[:length]
    return normalized + (" " * (length - len(normalized)))


def fmt_numeric_int(value: Optional[int], length: int) -> str:
    if value is None:
        value = 0
    if value < 0:
        raise ValueError("Negative numeric field not allowed without sign.")
    raw = str(int(value))
    if len(raw) > length:
        raise ValueError(f"Numeric field '{raw}' exceeds length {length}.")
    return "0" * (length - len(raw)) + raw


def fmt_amount_euros(value: Optional[Decimal], int_len: int, dec_len: int = 2) -> Tuple[str, str]:
    if value is None:
        return ("0" * int_len, "0" * dec_len)
    amount = _quantize_eur(_to_decimal(value))
    cents = int(amount * 100)
    if cents < 0:
        raise ValueError("Negative amounts require separate sign fields.")
    integer = cents // 100
    decimals = cents % 100
    return (fmt_numeric_int(integer, int_len), fmt_numeric_int(decimals, dec_len))


def fmt_sign(is_negative: bool) -> str:
    return "N" if is_negative else " "


def fmt_nif(value: Optional[str]) -> str:
    if not value:
        return "0" * 9
    cleaned = normalize_text(value).replace(" ", "")
    if len(cleaned) > 9:
        raise ValueError(f"NIF '{value}' exceeds 9 characters.")
    return cleaned.rjust(9, "0")


def fmt_nif_or_spaces(value: Optional[str]) -> str:
    if not value:
        return " " * 9
    return fmt_nif(value)


def _json_text(column, key: str):
    value = column[key]
    if hasattr(value, "as_string"):
        return value.as_string()
    if hasattr(value, "astext"):
        return value.astext
    return value


def _json_date(column, key: str):
    return cast(_json_text(column, key), Date)


def _put(buf: List[str], start: int, end: int, value: str) -> None:
    if len(value) != (end - start + 1):
        raise ValueError(f"Value length mismatch for positions {start}-{end}.")
    buf[start - 1 : end] = list(value)


def load_mapping_config(path: Optional[str]) -> MappingConfig:
    if not path:
        return MappingConfig()
    with open(path, "r", encoding="utf-8") as handle:
        raw = json.load(handle)
    concepts_raw = raw.get("concepts_to_clave_subclave", [])
    rules: List[ConceptMappingRule] = []
    for entry in concepts_raw:
        if not isinstance(entry, dict):
            continue
        clave = (entry.get("clave") or "").strip().upper()
        subclave = entry.get("subclave")
        subclave = subclave.strip().zfill(2) if isinstance(subclave, str) and subclave.strip() else None
        concepts = entry.get("concepts") or []
        rules.append(ConceptMappingRule(clave=clave, subclave=subclave, concepts=concepts))
    allowed_claves = [str(val).upper() for val in raw.get("allowed_claves", ["A", "L"])]
    allowed_subclaves = {
        str(key).upper(): [str(val).zfill(2) for val in values]
        for key, values in raw.get("allowed_subclaves_by_clave", {"L": ["01", "05", "24", "25", "26"]}).items()
    }
    allow_gastos = [str(val).upper() for val in raw.get("allow_gastos_deducibles_claves", ["A", "L.05", "L.10", "L.27"])]
    config = MappingConfig(
        concepts_to_clave_subclave=rules,
        ss_tax_concepts=raw.get("ss_tax_concepts", []),
        allowed_claves=allowed_claves,
        allowed_subclaves_by_clave=allowed_subclaves,
        gastos_deducibles_allocation=raw.get("gastos_deducibles_allocation", "clave_a"),
        allow_gastos_deducibles_claves=allow_gastos,
        include_in_kind=bool(raw.get("include_in_kind", False)),
        default_provincia=str(raw.get("default_provincia", "98")),
    )
    return config


def parse_date(value: str) -> date:
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value}")


def _resolve_period(year: Optional[int], date_from: Optional[str], date_to: Optional[str]) -> Tuple[date, date, int]:
    if year is not None:
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        return start, end, year
    if not date_from or not date_to:
        raise ValueError("Provide either --year or both --from/--to.")
    start = parse_date(date_from)
    end = parse_date(date_to)
    if end < start:
        raise ValueError("--to must be after --from.")
    return start, end, end.year


def _build_employee_name(employee: EmployeeInfo) -> str:
    parts = [employee.last_name]
    if employee.last_name2:
        parts.append(employee.last_name2)
    parts.append(employee.first_name)
    return " ".join(part for part in parts if part)


def _classify_clave_subclave(
    line: PayrollLine,
    concept_map: Dict[str, Tuple[str, Optional[str]]],
) -> Tuple[str, Optional[str]]:
    if line.is_sickpay:
        return "A", None
    key = normalize_concept_key(line.concept)
    return concept_map.get(key, ("A", None))


def _is_ss_tax(line: PayrollLine, ss_tax_set: set[str]) -> bool:
    if line.category != "deduccion":
        return False
    key = normalize_concept_key(line.concept)
    return key in ss_tax_set


def fetch_payroll_lines_for_cif_period(
    session: Session,
    company_cif: str,
    range_start: date,
    range_end: date,
) -> Iterable:
    period_start = _json_date(Payroll.periodo, "desde")
    period_end = _json_date(Payroll.periodo, "hasta")
    payroll_start = func.coalesce(period_start, period_end)

    eligible_payrolls = (
        session.query(Payroll.id.label("payroll_id"))
        .join(Employee, Payroll.employee_id == Employee.id)
        .join(EmployeePeriod, EmployeePeriod.employee_id == Employee.id)
        .join(ClientLocation, EmployeePeriod.location_id == ClientLocation.id)
        .join(Client, ClientLocation.company_id == Client.id)
        .filter(Client.cif == company_cif)
        .filter(period_end.isnot(None))
        .filter(payroll_start <= range_end)
        .filter(period_end >= range_start)
        .filter(EmployeePeriod.period_begin_date <= period_end)
        .filter(
            or_(EmployeePeriod.period_end_date.is_(None), EmployeePeriod.period_end_date >= payroll_start)
        )
        .distinct()
        .subquery()
    )

    return (
        session.query(
            Payroll.id.label("payroll_id"),
            Payroll.tipo_irpf.label("tipo_irpf"),
            _json_text(Payroll.periodo, "desde").label("periodo_desde"),
            _json_text(Payroll.periodo, "hasta").label("periodo_hasta"),
            Employee.id.label("employee_id"),
            Employee.identity_card_number.label("nif"),
            Employee.first_name.label("first_name"),
            Employee.last_name.label("last_name"),
            Employee.last_name2.label("last_name2"),
            Employee.birth_date.label("birth_date"),
            Employee.address.label("address"),
            PayrollLine.id.label("line_id"),
            PayrollLine.category.label("category"),
            PayrollLine.concept.label("concept"),
            PayrollLine.amount.label("amount"),
            PayrollLine.is_taxable_income.label("is_taxable_income"),
            PayrollLine.is_sickpay.label("is_sickpay"),
            PayrollLine.is_in_kind.label("is_in_kind"),
        )
        .select_from(PayrollLine)
        .join(Payroll, PayrollLine.payroll_id == Payroll.id)
        .join(Employee, Payroll.employee_id == Employee.id)
        .filter(Payroll.id.in_(select(eligible_payrolls.c.payroll_id)))
        .all()
    )


def _allocate_gastos_deducibles(
    employee_groups: Dict[GroupKey, Aggregation],
    employee_totals: Dict[int, Decimal],
    allocation_mode: str,
) -> None:
    if allocation_mode == "clave_a":
        for group_key, agg in employee_groups.items():
            if group_key.clave == "A" and group_key.subclave is None:
                agg.gastos_deducibles = _quantize_eur(employee_totals.get(group_key.employee_id, _DECIMAL_ZERO))
        return

    if allocation_mode != "proportional":
        raise ValueError(f"Unsupported gastos_deducibles_allocation: {allocation_mode}")

    grouped_by_employee: Dict[int, List[Tuple[GroupKey, Aggregation]]] = {}
    for group_key, agg in employee_groups.items():
        grouped_by_employee.setdefault(group_key.employee_id, []).append((group_key, agg))

    for employee_id, group_items in grouped_by_employee.items():
        total_gastos = employee_totals.get(employee_id, _DECIMAL_ZERO)
        if total_gastos == _DECIMAL_ZERO:
            continue
        base_total = sum((item.devengo_base_no_it for _, item in group_items), _DECIMAL_ZERO)
        if base_total == _DECIMAL_ZERO:
            for group_key, agg in group_items:
                if group_key.clave == "A" and group_key.subclave is None:
                    agg.gastos_deducibles = _quantize_eur(total_gastos)
            continue

        remaining = _quantize_eur(total_gastos)
        for idx, (group_key, agg) in enumerate(sorted(group_items, key=lambda x: (x[0].clave, x[0].subclave or ""))):
            if idx == len(group_items) - 1:
                agg.gastos_deducibles = remaining
            else:
                share = (total_gastos * agg.devengo_base_no_it) / base_total
                share = _quantize_eur(share)
                agg.gastos_deducibles = share
                remaining -= share


def build_perceptor_inputs(
    session: Session,
    company_cif: str,
    range_start: date,
    range_end: date,
    config: MappingConfig,
) -> List[PerceptorRecordInput]:
    concept_map = config.build_concept_map()
    ss_tax_set = {normalize_concept_key(value) for value in config.ss_tax_concepts}

    rows = fetch_payroll_lines_for_cif_period(session, company_cif, range_start, range_end)

    employee_info: Dict[int, EmployeeInfo] = {}
    groups: Dict[GroupKey, Aggregation] = {}
    taxable_base_by_payroll: Dict[Tuple[int, str, Optional[str], int, bool], Decimal] = {}
    payroll_tipo_irpf: Dict[int, Decimal] = {}
    ss_tax_by_employee: Dict[int, Decimal] = {}

    for row in rows:
        payroll_id = row.payroll_id
        employee_id = row.employee_id

        if employee_id not in employee_info:
            employee_info[employee_id] = EmployeeInfo(
                employee_id=employee_id,
                nif=row.nif or "",
                first_name=row.first_name or "",
                last_name=row.last_name or "",
                last_name2=row.last_name2,
                birth_date=row.birth_date,
                address=row.address,
            )

        payroll_tipo_irpf[payroll_id] = _to_decimal(row.tipo_irpf)

        if row.is_in_kind and not config.include_in_kind:
            continue

        category = (row.category or "").lower()

        dummy_line = PayrollLine(
            id=row.line_id,
            payroll_id=payroll_id,
            category=category,
            concept=row.concept,
            raw_concept=None,
            amount=row.amount,
            is_taxable_income=row.is_taxable_income,
            is_taxable_ss=False,
            is_sickpay=row.is_sickpay,
            is_in_kind=row.is_in_kind,
            is_pay_advance=False,
            is_seizure=False,
        )

        clave, subclave = _classify_clave_subclave(dummy_line, concept_map)
        group_key = GroupKey(employee_id=employee_id, clave=clave, subclave=subclave)
        agg = groups.setdefault(group_key, Aggregation())

        amount = _to_decimal(row.amount)

        if category == "devengo":
            if row.is_sickpay:
                agg.percepcion_it += amount
            else:
                agg.percepcion_no_it += amount
                agg.devengo_base_no_it += amount

            if row.is_taxable_income:
                key = (employee_id, clave, subclave, payroll_id, bool(row.is_sickpay))
                taxable_base_by_payroll[key] = taxable_base_by_payroll.get(key, _DECIMAL_ZERO) + amount

        elif category == "deduccion":
            if _is_ss_tax(dummy_line, ss_tax_set):
                ss_tax_by_employee[employee_id] = ss_tax_by_employee.get(employee_id, _DECIMAL_ZERO) + amount

    for (employee_id, clave, subclave, payroll_id, is_sickpay), base in taxable_base_by_payroll.items():
        tipo_irpf = payroll_tipo_irpf.get(payroll_id, _DECIMAL_ZERO)
        retention = _quantize_eur(base * tipo_irpf / Decimal("100"))
        group_key = GroupKey(employee_id=employee_id, clave=clave, subclave=subclave)
        agg = groups.setdefault(group_key, Aggregation())
        if is_sickpay:
            agg.retencion_it += retention
        else:
            agg.retencion_no_it += retention

    _allocate_gastos_deducibles(groups, ss_tax_by_employee, config.gastos_deducibles_allocation)

    perceptors: List[PerceptorRecordInput] = []
    for group_key, agg in groups.items():
        has_amounts = any(
            value > _DECIMAL_ZERO
            for value in [
                agg.percepcion_no_it,
                agg.percepcion_it,
                agg.retencion_no_it,
                agg.retencion_it,
            ]
        )
        if not has_amounts:
            continue

        employee = employee_info[group_key.employee_id]
        name = _build_employee_name(employee)
        datos_adicionales = AdditionalData(
            anio_nacimiento=employee.birth_date.year if employee.birth_date else None,
            situacion_familiar=3,
            discapacidad=0,
            contrato_relacion=1 if group_key.clave == "A" else None,
            movilidad_geografica=False if group_key.clave == "A" else None,
            gastos_deducibles=agg.gastos_deducibles,
            comunicacion_prestamo_vivienda=False,
        )

        dinerario_no_it = None
        if agg.percepcion_no_it > _DECIMAL_ZERO or agg.retencion_no_it > _DECIMAL_ZERO:
            dinerario_no_it = IngresoDinerario(
                base=_quantize_eur(agg.percepcion_no_it),
                retencion=_quantize_eur(agg.retencion_no_it),
            )

        incapacidad = None
        if agg.percepcion_it > _DECIMAL_ZERO or agg.retencion_it > _DECIMAL_ZERO:
            incapacidad = Incapacidad(
                dineraria=IngresoDinerario(
                    base=_quantize_eur(agg.percepcion_it),
                    retencion=_quantize_eur(agg.retencion_it),
                )
            )

        perceptors.append(
            PerceptorRecordInput(
                nif_perceptor=employee.nif,
                nombre_perceptor=name,
                provincia=str(config.default_provincia).zfill(2),
                clave=group_key.clave,
                subclave=group_key.subclave,
                dinerario_no_incapacidad=dinerario_no_it,
                incapacidad=incapacidad,
                ceuta_melilla_flag=0,
                ejercicio_devengo=0,
                excesos_acciones_emergentes=0,
                datos_adicionales=datos_adicionales,
            )
        )

    perceptors.sort(key=lambda p: (normalize_text(p.nif_perceptor), p.clave, p.subclave or ""))
    return perceptors


def compute_totals(perceptors: Sequence[PerceptorRecordInput]) -> Tuple[Decimal, Decimal]:
    total_percepciones = _DECIMAL_ZERO
    total_ret_ing = _DECIMAL_ZERO

    for p in perceptors:
        if p.dinerario_no_incapacidad:
            sign = -1 if p.dinerario_no_incapacidad.sign == "N" else 1
            total_percepciones += sign * _to_decimal(p.dinerario_no_incapacidad.base)
            total_ret_ing += _to_decimal(p.dinerario_no_incapacidad.retencion)

        if p.especie_no_incapacidad:
            sign = -1 if p.especie_no_incapacidad.sign == "N" else 1
            total_percepciones += sign * _to_decimal(p.especie_no_incapacidad.base)
            total_ret_ing += _to_decimal(p.especie_no_incapacidad.ingreso_cuenta_efectuado)

        if p.incapacidad and p.incapacidad.dineraria:
            sign = -1 if p.incapacidad.dineraria.sign == "N" else 1
            total_percepciones += sign * _to_decimal(p.incapacidad.dineraria.base)
            total_ret_ing += _to_decimal(p.incapacidad.dineraria.retencion)

        if p.incapacidad and p.incapacidad.especie:
            sign = -1 if p.incapacidad.especie.sign == "N" else 1
            total_percepciones += sign * _to_decimal(p.incapacidad.especie.base)
            total_ret_ing += _to_decimal(p.incapacidad.especie.ingreso_cuenta_efectuado)

    return (_quantize_eur(total_percepciones), _quantize_eur(total_ret_ing))


def _build_datos_adicionales(buf: List[str], p: PerceptorRecordInput, config: MappingConfig) -> None:
    _put(buf, 153, 254, "0" * 102)
    _put(buf, 158, 166, " " * 9)

    datos = p.datos_adicionales
    if not datos:
        return

    year = datos.anio_nacimiento or 0
    _put(buf, 153, 156, fmt_numeric_int(year, 4))
    situacion = datos.situacion_familiar if datos.situacion_familiar is not None else 3
    _put(buf, 157, 157, fmt_numeric_int(situacion, 1))

    if datos.nif_conyuge:
        _put(buf, 158, 166, fmt_nif(datos.nif_conyuge))

    discapacidad = datos.discapacidad if datos.discapacidad is not None else 0
    _put(buf, 167, 167, fmt_numeric_int(discapacidad, 1))

    if p.clave == "A":
        contrato = datos.contrato_relacion if datos.contrato_relacion is not None else 1
        _put(buf, 168, 168, fmt_numeric_int(contrato, 1))
        movilidad = 1 if datos.movilidad_geografica else 0
        _put(buf, 170, 170, fmt_numeric_int(movilidad, 1))

    gastos_allowed = False
    if p.clave == "A":
        gastos_allowed = True
    elif p.clave == "L" and p.subclave:
        token = f"L.{p.subclave}".upper()
        gastos_allowed = token in config.allow_gastos_deducibles_claves

    if gastos_allowed and datos.gastos_deducibles is not None:
        ent, dec = fmt_amount_euros(datos.gastos_deducibles, 11)
        _put(buf, 184, 194, ent)
        _put(buf, 195, 196, dec)

    prestamo = 1 if datos.comunicacion_prestamo_vivienda else 0
    _put(buf, 254, 254, fmt_numeric_int(prestamo, 1))


def _build_incapacidad_dineraria(buf: List[str], p: PerceptorRecordInput) -> None:
    inc = p.incapacidad.dineraria if p.incapacidad else None
    if inc and (inc.base > _DECIMAL_ZERO or inc.retencion > _DECIMAL_ZERO):
        sign = fmt_sign(inc.sign == "N") if inc.sign else " "
        ent, dec = fmt_amount_euros(inc.base, 11)
        ret_ent, ret_dec = fmt_amount_euros(inc.retencion, 11)
    else:
        sign = " "
        ent = "0" * 11
        dec = "00"
        ret_ent = "0" * 11
        ret_dec = "00"
    _put(buf, 255, 255, sign)
    _put(buf, 256, 266, ent)
    _put(buf, 267, 268, dec)
    _put(buf, 269, 279, ret_ent)
    _put(buf, 280, 281, ret_dec)


def _build_incapacidad_especie(buf: List[str], p: PerceptorRecordInput) -> None:
    inc = p.incapacidad.especie if p.incapacidad else None
    if inc and (
        inc.base > _DECIMAL_ZERO
        or inc.ingreso_cuenta_efectuado > _DECIMAL_ZERO
        or inc.ingreso_cuenta_repercutido > _DECIMAL_ZERO
    ):
        sign = fmt_sign(inc.sign == "N") if inc.sign else " "
        ent, dec = fmt_amount_euros(inc.base, 11)
        ing_ent, ing_dec = fmt_amount_euros(inc.ingreso_cuenta_efectuado, 11)
        rep_ent, rep_dec = fmt_amount_euros(inc.ingreso_cuenta_repercutido, 11)
    else:
        sign = " "
        ent = "0" * 11
        dec = "00"
        ing_ent = "0" * 11
        ing_dec = "00"
        rep_ent = "0" * 11
        rep_dec = "00"
    _put(buf, 282, 282, sign)
    _put(buf, 283, 293, ent)
    _put(buf, 294, 295, dec)
    _put(buf, 296, 306, ing_ent)
    _put(buf, 307, 308, ing_dec)
    _put(buf, 309, 319, rep_ent)
    _put(buf, 320, 321, rep_dec)


def _build_foral_split(buf: List[str], p: PerceptorRecordInput) -> None:
    _put(buf, 323, 387, "0" * 65)
    if not p.foral_split:
        return
    splits = p.foral_split
    ent, _ = fmt_amount_euros(splits.hacienda_estatal, 13)
    _put(buf, 323, 335, ent)
    ent, _ = fmt_amount_euros(splits.navarra, 13)
    _put(buf, 336, 348, ent)
    ent, _ = fmt_amount_euros(splits.araba, 13)
    _put(buf, 349, 361, ent)
    ent, _ = fmt_amount_euros(splits.gipuzkoa, 13)
    _put(buf, 362, 374, ent)
    ent, _ = fmt_amount_euros(splits.bizkaia, 13)
    _put(buf, 375, 387, ent)


def build_type2_record(decl: Declarant190, p: PerceptorRecordInput, config: MappingConfig) -> str:
    buf = [" "] * 500

    _put(buf, 1, 1, "2")
    _put(buf, 2, 4, "190")
    _put(buf, 5, 8, fmt_numeric_int(decl.ejercicio, 4))
    _put(buf, 9, 17, fmt_nif(decl.nif))
    _put(buf, 18, 26, fmt_nif(p.nif_perceptor))
    _put(buf, 27, 35, fmt_nif_or_spaces(p.nif_representante))
    _put(buf, 36, 75, fmt_alpha(p.nombre_perceptor, 40))
    _put(buf, 76, 77, fmt_numeric_int(int(p.provincia), 2))
    _put(buf, 78, 78, p.clave)

    if p.clave in {"B", "C", "E", "F", "G", "H", "I", "K", "L"}:
        subclave_val = p.subclave or ""
        _put(buf, 79, 80, fmt_numeric_int(int(subclave_val or "0"), 2))
    else:
        _put(buf, 79, 80, "  ")

    dn = p.dinerario_no_incapacidad
    if dn and (dn.base > _DECIMAL_ZERO or dn.retencion > _DECIMAL_ZERO):
        sign = fmt_sign(dn.sign == "N") if dn.sign else " "
        ent, dec = fmt_amount_euros(dn.base, 11)
        ret_ent, ret_dec = fmt_amount_euros(dn.retencion, 11)
    else:
        sign = " "
        ent = "0" * 11
        dec = "00"
        ret_ent = "0" * 11
        ret_dec = "00"
    _put(buf, 81, 81, sign)
    _put(buf, 82, 92, ent)
    _put(buf, 93, 94, dec)
    _put(buf, 95, 105, ret_ent)
    _put(buf, 106, 107, ret_dec)

    es = p.especie_no_incapacidad
    if es and (
        es.base > _DECIMAL_ZERO
        or es.ingreso_cuenta_efectuado > _DECIMAL_ZERO
        or es.ingreso_cuenta_repercutido > _DECIMAL_ZERO
    ):
        sign = fmt_sign(es.sign == "N") if es.sign else " "
        ent, dec = fmt_amount_euros(es.base, 11)
        ing_ent, ing_dec = fmt_amount_euros(es.ingreso_cuenta_efectuado, 11)
        rep_ent, rep_dec = fmt_amount_euros(es.ingreso_cuenta_repercutido, 11)
    else:
        sign = " "
        ent = "0" * 11
        dec = "00"
        ing_ent = "0" * 11
        ing_dec = "00"
        rep_ent = "0" * 11
        rep_dec = "00"
    _put(buf, 108, 108, sign)
    _put(buf, 109, 119, ent)
    _put(buf, 120, 121, dec)
    _put(buf, 122, 132, ing_ent)
    _put(buf, 133, 134, ing_dec)
    _put(buf, 135, 145, rep_ent)
    _put(buf, 146, 147, rep_dec)

    _put(buf, 148, 151, fmt_numeric_int(p.ejercicio_devengo or 0, 4))
    _put(buf, 152, 152, fmt_numeric_int(p.ceuta_melilla_flag or 0, 1))

    _build_datos_adicionales(buf, p, config)
    _build_incapacidad_dineraria(buf, p)
    _build_incapacidad_especie(buf, p)

    _put(buf, 322, 322, fmt_numeric_int(p.complemento_ayuda_infancia or 0, 1))
    _build_foral_split(buf, p)
    _put(buf, 388, 388, fmt_numeric_int(p.excesos_acciones_emergentes or 0, 1))

    record = "".join(buf)
    if len(record) != 500:
        raise ValueError("Type 2 record length mismatch.")
    return record


def build_type1_record(decl: Declarant190, perceptors: Sequence[PerceptorRecordInput]) -> str:
    buf = [" "] * 500

    _put(buf, 1, 1, "1")
    _put(buf, 2, 4, "190")
    _put(buf, 5, 8, fmt_numeric_int(decl.ejercicio, 4))
    _put(buf, 9, 17, fmt_nif(decl.nif))
    _put(buf, 18, 57, fmt_alpha(decl.nombre_razon_social, 40))
    _put(buf, 58, 58, "T")

    telefono = "".join(ch for ch in (decl.contacto_telefono or "") if ch.isdigit())
    telefono_val = int(telefono) if telefono else 0
    _put(buf, 59, 67, fmt_numeric_int(telefono_val, 9))
    _put(buf, 68, 107, fmt_alpha(decl.contacto_nombre or "", 40))

    numero_ident = int(decl.numero_identificativo) if decl.numero_identificativo else 0
    _put(buf, 108, 120, fmt_numeric_int(numero_ident, 13))

    if decl.tipo_declaracion == "C":
        _put(buf, 121, 121, "C")
        _put(buf, 122, 122, " ")
    elif decl.tipo_declaracion == "S":
        _put(buf, 121, 121, " ")
        _put(buf, 122, 122, "S")
    else:
        _put(buf, 121, 121, " ")
        _put(buf, 122, 122, " ")

    numero_prev = int(decl.id_declaracion_anterior) if decl.id_declaracion_anterior else 0
    _put(buf, 123, 135, fmt_numeric_int(numero_prev, 13))

    n_perceptors = len(perceptors)
    _put(buf, 136, 144, fmt_numeric_int(n_perceptors, 9))

    total_per, total_ret = compute_totals(perceptors)
    sign = fmt_sign(total_per < 0)
    abs_per = abs(total_per)
    per_ent, per_dec = fmt_amount_euros(abs_per, 13)
    _put(buf, 145, 145, sign)
    _put(buf, 146, 158, per_ent)
    _put(buf, 159, 160, per_dec)

    ret_ent, ret_dec = fmt_amount_euros(total_ret, 13)
    _put(buf, 161, 173, ret_ent)
    _put(buf, 174, 175, ret_dec)

    _put(buf, 176, 225, fmt_alpha(decl.email_contacto or "", 50))

    record = "".join(buf)
    if len(record) != 500:
        raise ValueError("Type 1 record length mismatch.")
    return record


def validate_perceptors(perceptors: Sequence[PerceptorRecordInput], config: MappingConfig) -> List[str]:
    errors: List[str] = []
    if not perceptors:
        errors.append("No perceptor records generated.")
        return errors

    allowed_claves = set(config.allowed_claves)
    allowed_subclaves_by_clave = {
        key: {str(v).zfill(2) for v in values} for key, values in config.allowed_subclaves_by_clave.items()
    }

    for idx, p in enumerate(perceptors, start=1):
        if p.clave not in allowed_claves:
            errors.append(f"Perceptor {idx}: clave '{p.clave}' not allowed.")

        needs_subclave = p.clave in {"B", "C", "E", "F", "G", "H", "I", "K", "L"}
        if needs_subclave and not p.subclave:
            errors.append(f"Perceptor {idx}: subclave required for clave '{p.clave}'.")
        if not needs_subclave and p.subclave:
            errors.append(f"Perceptor {idx}: subclave should be empty for clave '{p.clave}'.")

        if p.clave in allowed_subclaves_by_clave:
            if p.subclave and p.subclave not in allowed_subclaves_by_clave[p.clave]:
                errors.append(
                    f"Perceptor {idx}: subclave '{p.subclave}' not allowed for clave '{p.clave}'."
                )

        if p.incapacidad and (p.incapacidad.dineraria or p.incapacidad.especie):
            if p.clave != "A":
                errors.append(f"Perceptor {idx}: incapacidad amounts only allowed for clave A.")

        def _validate_ingreso_dinerario(component: IngresoDinerario, label: str) -> None:
            if component.base < _DECIMAL_ZERO or component.retencion < _DECIMAL_ZERO:
                errors.append(f"Perceptor {idx}: negative amounts in {label}.")
            if component.sign == "N":
                if component.base <= _DECIMAL_ZERO:
                    errors.append(f"Perceptor {idx}: sign N without base in {label}.")
                if component.retencion > _DECIMAL_ZERO:
                    errors.append(f"Perceptor {idx}: sign N requires zero retencion in {label}.")

        def _validate_ingreso_especie(component: IngresoEspecie, label: str) -> None:
            if (
                component.base < _DECIMAL_ZERO
                or component.ingreso_cuenta_efectuado < _DECIMAL_ZERO
                or component.ingreso_cuenta_repercutido < _DECIMAL_ZERO
            ):
                errors.append(f"Perceptor {idx}: negative amounts in {label}.")
            if component.sign == "N":
                if component.base <= _DECIMAL_ZERO:
                    errors.append(f"Perceptor {idx}: sign N without base in {label}.")
                if (
                    component.ingreso_cuenta_efectuado > _DECIMAL_ZERO
                    or component.ingreso_cuenta_repercutido > _DECIMAL_ZERO
                ):
                    errors.append(f"Perceptor {idx}: sign N requires zero ingresos in {label}.")

        if p.dinerario_no_incapacidad:
            _validate_ingreso_dinerario(p.dinerario_no_incapacidad, "dinerario_no_incapacidad")
        if p.especie_no_incapacidad:
            _validate_ingreso_especie(p.especie_no_incapacidad, "especie_no_incapacidad")
        if p.incapacidad:
            if p.incapacidad.dineraria:
                _validate_ingreso_dinerario(p.incapacidad.dineraria, "incapacidad_dineraria")
            if p.incapacidad.especie:
                _validate_ingreso_especie(p.incapacidad.especie, "incapacidad_especie")

        if p.foral_split:
            split_total = sum(
                [
                    p.foral_split.hacienda_estatal,
                    p.foral_split.navarra,
                    p.foral_split.araba,
                    p.foral_split.gipuzkoa,
                    p.foral_split.bizkaia,
                ],
                _DECIMAL_ZERO,
            )
            ret_total = _DECIMAL_ZERO
            if p.dinerario_no_incapacidad:
                ret_total += p.dinerario_no_incapacidad.retencion
            if p.especie_no_incapacidad:
                ret_total += p.especie_no_incapacidad.ingreso_cuenta_efectuado
            if split_total != _quantize_eur(ret_total):
                errors.append(f"Perceptor {idx}: foral split does not match retenciones/ingresos.")

        total_non_zero = False
        if p.dinerario_no_incapacidad:
            if p.dinerario_no_incapacidad.base > _DECIMAL_ZERO or p.dinerario_no_incapacidad.retencion > _DECIMAL_ZERO:
                total_non_zero = True
        if p.especie_no_incapacidad:
            if (
                p.especie_no_incapacidad.base > _DECIMAL_ZERO
                or p.especie_no_incapacidad.ingreso_cuenta_efectuado > _DECIMAL_ZERO
                or p.especie_no_incapacidad.ingreso_cuenta_repercutido > _DECIMAL_ZERO
            ):
                total_non_zero = True
        if p.incapacidad:
            if p.incapacidad.dineraria:
                if (
                    p.incapacidad.dineraria.base > _DECIMAL_ZERO
                    or p.incapacidad.dineraria.retencion > _DECIMAL_ZERO
                ):
                    total_non_zero = True
            if p.incapacidad.especie:
                if (
                    p.incapacidad.especie.base > _DECIMAL_ZERO
                    or p.incapacidad.especie.ingreso_cuenta_efectuado > _DECIMAL_ZERO
                    or p.incapacidad.especie.ingreso_cuenta_repercutido > _DECIMAL_ZERO
                ):
                    total_non_zero = True

        if not total_non_zero:
            errors.append(f"Perceptor {idx}: record has no monetary amounts.")

    return errors


def generate_190_file(decl: Declarant190, config: MappingConfig) -> bytes:
    perceptors = decl.percepciones
    errors = validate_perceptors(perceptors, config)
    if errors:
        raise ValueError("Validation errors:\n" + "\n".join(errors))

    header = build_type1_record(decl, perceptors)
    records = [build_type2_record(decl, p, config) for p in perceptors]
    file_text = "\r\n".join([header] + records) + "\r\n"
    return file_text.encode("iso-8859-1")


def generate_190_xlsx(decl: Declarant190, config: MappingConfig, output_path: str) -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to generate XLSX output. Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo 190"

    ws.append(["DECLARANTE"])
    ws.append(["Ejercicio", decl.ejercicio])
    ws.append(["Modelo", decl.modelo])
    ws.append(["NIF", decl.nif])
    ws.append(["Nombre/Razon social", decl.nombre_razon_social])
    ws.append(["Tipo declaracion", decl.tipo_declaracion])
    ws.append(["Numero identificativo", decl.numero_identificativo or ""])
    ws.append(["Id declaracion anterior", decl.id_declaracion_anterior or ""])
    ws.append(["Contacto nombre", decl.contacto_nombre or ""])
    ws.append(["Contacto telefono", decl.contacto_telefono or ""])
    ws.append(["Email contacto", decl.email_contacto or ""])
    ws.append([])

    header = [
        "NIF perceptor",
        "Nombre perceptor",
        "Provincia",
        "Clave",
        "Subclave",
        "Percepcion no IT",
        "Retencion no IT",
        "Gastos deducibles",
        "Percepcion IT",
        "Retencion IT",
    ]
    ws.append(header)

    currency_format = "#,##0.00"
    for row_idx, p in enumerate(decl.percepciones, start=ws.max_row + 1):
        dinerario = p.dinerario_no_incapacidad
        incap = p.incapacidad.dineraria if p.incapacidad else None
        gastos = p.datos_adicionales.gastos_deducibles if p.datos_adicionales else _DECIMAL_ZERO
        ws.append(
            [
                p.nif_perceptor,
                p.nombre_perceptor,
                p.provincia,
                p.clave,
                p.subclave or "",
                _as_float(dinerario.base) if dinerario else 0.0,
                _as_float(dinerario.retencion) if dinerario else 0.0,
                _as_float(gastos),
                _as_float(incap.base) if incap else 0.0,
                _as_float(incap.retencion) if incap else 0.0,
            ]
        )
        for col in range(6, 11):
            ws.cell(row=row_idx, column=col).number_format = currency_format

    wb.save(output_path)


def build_declarant_from_cif(
    session: Session,
    company_cif: str,
    ejercicio: int,
    percepciones: List[PerceptorRecordInput],
    contacto_nombre: Optional[str],
    contacto_telefono: Optional[str],
    email_contacto: Optional[str],
    numero_identificativo: Optional[str],
    tipo_declaracion: str,
    id_declaracion_anterior: Optional[str],
) -> Declarant190:
    client = session.query(Client).filter(Client.cif == company_cif).first()
    if not client:
        raise ValueError(f"Client with CIF '{company_cif}' not found.")
    return Declarant190(
        ejercicio=ejercicio,
        modelo="190",
        nif=company_cif,
        nombre_razon_social=client.name or "",
        contacto_telefono=contacto_telefono,
        contacto_nombre=contacto_nombre,
        email_contacto=email_contacto,
        numero_identificativo=numero_identificativo,
        tipo_declaracion=tipo_declaracion,
        id_declaracion_anterior=id_declaracion_anterior,
        percepciones=percepciones,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate Modelo 190 file from payroll data.")
    parser.add_argument("--cif", required=True, help="Company CIF (declarante).")
    parser.add_argument("--year", type=int, help="Calendar year to report (YYYY).")
    parser.add_argument("--from", dest="date_from", help="Custom period start YYYY-MM-DD.")
    parser.add_argument("--to", dest="date_to", help="Custom period end YYYY-MM-DD.")
    parser.add_argument("--ejercicio", type=int, help="Ejercicio for the declaration (defaults to year/end year).")
    parser.add_argument("--mapping", help="Path to JSON mapping config.")
    parser.add_argument("--out", required=True, help="Output .txt path.")
    parser.add_argument("--xlsx", help="Optional XLSX output path.")
    parser.add_argument("--contact-name", help="Contact person full name.")
    parser.add_argument("--contact-phone", help="Contact phone digits.")
    parser.add_argument("--contact-email", help="Contact email.")
    parser.add_argument(
        "--tipo-declaracion",
        choices=["N", "C", "S"],
        default="N",
        help="Tipo declaracion: N normal, C complementaria, S sustitutiva.",
    )
    parser.add_argument("--numero-identificativo", help="AEAT declaration ID (13 digits).")
    parser.add_argument("--id-declaracion-anterior", help="Previous AEAT declaration ID (13 digits).")
    parser.add_argument("--echo-sql", action="store_true", help="Enable SQL echo.")

    args = parser.parse_args()

    config = load_mapping_config(args.mapping)
    range_start, range_end, inferred_year = _resolve_period(args.year, args.date_from, args.date_to)
    ejercicio = args.ejercicio or inferred_year

    session = get_session(echo=args.echo_sql)
    try:
        perceptors = build_perceptor_inputs(session, args.cif, range_start, range_end, config)
        decl = build_declarant_from_cif(
            session,
            args.cif,
            ejercicio,
            perceptors,
            args.contact_name,
            args.contact_phone,
            args.contact_email,
            args.numero_identificativo,
            args.tipo_declaracion,
            args.id_declaracion_anterior,
        )
        output_bytes = generate_190_file(decl, config)
        with open(args.out, "wb") as handle:
            handle.write(output_bytes)
        if args.xlsx:
            generate_190_xlsx(decl, config, args.xlsx)
    finally:
        session.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())


__all__ = [
    "MappingConfig",
    "ConceptMappingRule",
    "Declarant190",
    "PerceptorRecordInput",
    "IngresoDinerario",
    "IngresoEspecie",
    "Incapacidad",
    "ForalSplit",
    "AdditionalData",
    "fetch_payroll_lines_for_cif_period",
    "build_perceptor_inputs",
    "build_type1_record",
    "build_type2_record",
    "generate_190_file",
    "generate_190_xlsx",
]
